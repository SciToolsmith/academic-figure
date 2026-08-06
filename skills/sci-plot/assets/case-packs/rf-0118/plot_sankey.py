from pathlib import Path
from collections import OrderedDict
import colorsys
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon, Rectangle
from matplotlib.colors import to_rgb


def read_xlsx(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(iterator)]
    records = [dict(zip(headers, row)) for row in iterator if any(value is not None for value in row)]
    workbook.close()
    return records, headers


def require_columns(headers, columns, source):
    missing = [name for name in columns if name not in headers]
    if missing:
        raise ValueError(f"{source} is missing required columns: {', '.join(missing)}")


def unique(values):
    return list(OrderedDict.fromkeys(values))


def totals_by(records, column, weight="_weight"):
    totals = OrderedDict()
    for row in records:
        key = row[column]
        totals[key] = totals.get(key, 0.0) + float(row[weight])
    return totals


def palette(size, start):
    return [colorsys.hls_to_rgb((start + i * 0.61803398875) % 1.0, 0.56, 0.52) for i in range(size)]


def adjust_positions(values, minimum, lower=0.025, upper=0.975):
    values = np.asarray(values, dtype=float)
    order = np.argsort(values)
    placed = np.clip(values[order], lower, upper)
    for i in range(1, len(placed)):
        placed[i] = max(placed[i], placed[i - 1] + minimum)
    if len(placed) and placed[-1] > upper:
        placed -= placed[-1] - upper
        for i in range(len(placed) - 2, -1, -1):
            placed[i] = min(placed[i], placed[i + 1] - minimum)
    if len(placed) and placed[0] < lower:
        placed += lower - placed[0]
    result = np.empty_like(placed)
    result[order] = placed
    return result


def build_flow(records, stages, orders, weight_column=None):
    data = []
    for record in records:
        row = dict(record)
        weight = float(row[weight_column]) if weight_column else 1.0
        if not np.isfinite(weight) or weight <= 0:
            raise ValueError("All flow weights must be finite and positive")
        row["_weight"] = weight
        data.append(row)
    total = sum(row["_weight"] for row in data)
    stage_orders = []
    for column, requested in zip(stages, orders):
        observed = unique(str(row[column]) for row in data)
        ordered = [value for value in requested if value in observed]
        ordered.extend(value for value in observed if value not in ordered)
        stage_orders.append(ordered)
    maximum_nodes = max(map(len, stage_orders))
    gap = 0.006 if maximum_nodes > 24 else 0.012
    scale = (0.88 - gap * (maximum_nodes - 1)) / total
    nodes = []
    starts = [0.57, 0.03, 0.42, 0.77]
    for stage_index, (column, ordered) in enumerate(zip(stages, stage_orders)):
        totals = totals_by(data, column)
        colors = palette(len(ordered), starts[stage_index % len(starts)])
        occupied = total * scale + gap * (len(ordered) - 1)
        cursor = 0.5 + occupied / 2
        stage_nodes = []
        for index, name in enumerate(ordered):
            value = float(totals[name])
            stage_nodes.append({"name": name, "value": value, "y1": cursor, "y0": cursor - value * scale, "color": colors[index]})
            cursor -= value * scale + gap
        nodes.append(stage_nodes)
    links_by_gap = []
    for stage_index in range(len(stages) - 1):
        source_column = stages[stage_index]
        target_column = stages[stage_index + 1]
        grouped = OrderedDict()
        for row in data:
            key = (row[source_column], row[target_column])
            grouped[key] = grouped.get(key, 0.0) + row["_weight"]
        source_index = {node["name"]: i for i, node in enumerate(nodes[stage_index])}
        target_index = {node["name"]: i for i, node in enumerate(nodes[stage_index + 1])}
        links = [{"source": source, "target": target, "weight": weight, "source_index": source_index[source], "target_index": target_index[target]} for (source, target), weight in grouped.items()]
        source_cursor = {node["name"]: node["y1"] for node in nodes[stage_index]}
        for link in sorted(links, key=lambda item: (item["source_index"], item["target_index"])):
            height = link["weight"] * scale
            link["sy1"] = source_cursor[link["source"]]
            link["sy0"] = source_cursor[link["source"]] - height
            source_cursor[link["source"]] -= height
        target_cursor = {node["name"]: node["y1"] for node in nodes[stage_index + 1]}
        for link in sorted(links, key=lambda item: (item["target_index"], item["source_index"])):
            height = link["weight"] * scale
            link["ty1"] = target_cursor[link["target"]]
            link["ty0"] = target_cursor[link["target"]] - height
            target_cursor[link["target"]] -= height
        links_by_gap.append(links)
    return {"stages": stages, "nodes": nodes, "links": links_by_gap, "total": total}


def draw_ribbon(ax, x0, x1, link, source_color, target_color, segments=30):
    t = np.linspace(0, 1, segments + 1)
    smooth = t * t * (3 - 2 * t)
    xs = x0 + (x1 - x0) * t
    upper = link["sy1"] + (link["ty1"] - link["sy1"]) * smooth
    lower = link["sy0"] + (link["ty0"] - link["sy0"]) * smooth
    source = np.asarray(to_rgb(source_color))
    target = np.asarray(to_rgb(target_color))
    for i in range(segments):
        fraction = (i + 0.5) / segments
        color = source * (1 - fraction) + target * fraction
        polygon = Polygon([[xs[i], upper[i]], [xs[i + 1], upper[i + 1]], [xs[i + 1], lower[i + 1]], [xs[i], lower[i]]], closed=True, facecolor=color, edgecolor="none", alpha=0.52, rasterized=True)
        ax.add_patch(polygon)


def format_value(value, integer):
    return f"{int(round(value)):,}" if integer else f"{value:.1f}"


def draw_flow(ax, flow, headers, title, subtitle, integer):
    stage_count = len(flow["stages"])
    x_positions = np.arange(stage_count, dtype=float)
    node_width = 0.055
    node_lookup = [{node["name"]: node for node in stage} for stage in flow["nodes"]]
    for stage_index, links in enumerate(flow["links"]):
        for link in sorted(links, key=lambda item: item["weight"], reverse=True):
            draw_ribbon(ax, x_positions[stage_index] + node_width / 2, x_positions[stage_index + 1] - node_width / 2, link, node_lookup[stage_index][link["source"]]["color"], node_lookup[stage_index + 1][link["target"]]["color"])
    for stage_index, stage_nodes in enumerate(flow["nodes"]):
        centers = [(node["y0"] + node["y1"]) / 2 for node in stage_nodes]
        minimum = 0.026 if len(stage_nodes) > 20 else 0.032
        labels = adjust_positions(centers, minimum)
        for node, label_y, center_y in zip(stage_nodes, labels, centers):
            height = max(node["y1"] - node["y0"], 0.0011)
            center = (node["y0"] + node["y1"]) / 2
            ax.add_patch(Rectangle((x_positions[stage_index] - node_width / 2, center - height / 2), node_width, height, facecolor=node["color"], edgecolor="#FFFFFF", linewidth=0.45, zorder=4))
            label = f"{node['name']}  {format_value(node['value'], integer)}"
            if stage_index == 0:
                text_x = x_positions[stage_index] - node_width / 2 - 0.025
                horizontal = "right"
                anchor_x = x_positions[stage_index] - node_width / 2
            else:
                text_x = x_positions[stage_index] + node_width / 2 + 0.025
                horizontal = "left"
                anchor_x = x_positions[stage_index] + node_width / 2
            if abs(label_y - center_y) > 0.004:
                ax.plot([anchor_x, text_x], [center_y, label_y], color="#8B959D", linewidth=0.38, zorder=5)
            ax.text(text_x, label_y, label, ha=horizontal, va="center", fontsize=6.2 if len(stage_nodes) > 20 else 7.2, color="#17222B", zorder=6, bbox={"boxstyle": "round,pad=0.12", "facecolor": "#FFFFFF", "edgecolor": "none", "alpha": 0.78})
        ax.text(x_positions[stage_index], 1.016, headers[stage_index].upper(), ha="center", va="bottom", fontsize=8.1, weight="bold", color="#52616B")
    ax.text(0, 1.105, title, transform=ax.transAxes, ha="left", va="bottom", fontsize=14.5, weight="bold", color="#15252D")
    ax.text(0, 1.058, subtitle, transform=ax.transAxes, ha="left", va="bottom", fontsize=8.7, color="#53636C")
    ax.set_xlim(-0.48, stage_count - 1 + 0.82)
    ax.set_ylim(-0.03, 1.055)
    ax.axis("off")


base = Path(__file__).resolve().parent
data1, headers1 = read_xlsx(base / "data1.xlsx")
data2, headers2 = read_xlsx(base / "data2.xlsx")
require_columns(headers1, ["Timeline", "Lineages Distribution", "Geographic Location", "P.M.A.s"], "data1.xlsx")
require_columns(headers2, ["cluster", "mRNA1", "mRNA2", "Freq"], "data2.xlsx")
if any(row["Timeline"] is None or row["Lineages Distribution"] is None or row["Geographic Location"] is None for row in data1):
    raise ValueError("data1.xlsx contains missing values in the first three stages")
for row in data1:
    row["Timeline"] = str(row["Timeline"])
    row["Lineages Distribution"] = str(row["Lineages Distribution"])
    row["Geographic Location"] = str(row["Geographic Location"])
    row["P.M.A.s / status"] = "Outside China / not applicable" if row["P.M.A.s"] is None else str(row["P.M.A.s"])
for row in data2:
    if any(row[column] is None for column in ["cluster", "mRNA1", "mRNA2", "Freq"]):
        raise ValueError("data2.xlsx contains missing flow values")
    for column in ["cluster", "mRNA1", "mRNA2"]:
        row[column] = str(row[column])
timeline_order = sorted(unique(row["Timeline"] for row in data1), key=lambda value: int(value.split("-")[0]))
lineage_order = [value for value in ["Paraphyletic cluster", "Sublineage 8.7"] if value in set(row["Lineages Distribution"] for row in data1)]
geographic_totals = totals_by([dict(row, _weight=1.0) for row in data1], "Geographic Location")
geographic_order = [value for value, _ in sorted(geographic_totals.items(), key=lambda item: item[1], reverse=True) if value != "China"] + (["China"] if "China" in geographic_totals else [])
pma_totals = totals_by([dict(row, _weight=1.0) for row in data1], "P.M.A.s / status")
pma_order = (["Outside China / not applicable"] if "Outside China / not applicable" in pma_totals else []) + [value for value, _ in sorted(pma_totals.items(), key=lambda item: item[1], reverse=True) if value != "Outside China / not applicable"]
flow1 = build_flow(data1, ["Timeline", "Lineages Distribution", "Geographic Location", "P.M.A.s / status"], [timeline_order, lineage_order, geographic_order, pma_order])
weighted2 = [dict(row, _weight=float(row["Freq"])) for row in data2]
cluster_order = [value for value, _ in sorted(totals_by(weighted2, "cluster").items(), key=lambda item: item[1], reverse=True)]
mrna1_order = [value for value, _ in sorted(totals_by(weighted2, "mRNA1").items(), key=lambda item: item[1], reverse=True)]
mrna2_order = [value for value, _ in sorted(totals_by(weighted2, "mRNA2").items(), key=lambda item: item[1], reverse=True)]
flow2 = build_flow(data2, ["cluster", "mRNA1", "mRNA2"], [cluster_order, mrna1_order, mrna2_order], "Freq")
fig, axes = plt.subplots(1, 2, figsize=(19, 11.5), gridspec_kw={"width_ratios": [1.16, 0.84]})
fig.patch.set_facecolor("#F8F7F3")
for axis in axes:
    axis.set_facecolor("#F8F7F3")
draw_flow(axes[0], flow1, ["Time window", "Lineage", "Country", "P.M.A.s / status"], "A  Isolate distribution across time, lineage and geography", f"data1.xlsx · n = {len(data1):,} records · ribbon width = record count", True)
sum_freq = sum(float(row["Freq"]) for row in data2)
draw_flow(axes[1], flow2, ["Cluster", "mRNA 1", "mRNA 2"], "B  Weighted cluster–mRNA associations", f"data2.xlsx · {len(data2):,} rows · Σ Freq = {sum_freq:.2f} · ribbon width = Freq", False)
missing_pma = sum(row["P.M.A.s"] is None for row in data1)
fig.suptitle("Two independent alluvial datasets", x=0.055, y=0.985, ha="left", fontsize=20, weight="bold", color="#10242C")
fig.text(0.055, 0.953, "Panels are intentionally separate: the workbooks contain no shared record-level key and use different flow measures.", ha="left", va="top", fontsize=10.2, color="#53636C")
fig.text(0.055, 0.018, f"Blank P.M.A.s values are retained as “Outside China / not applicable” (n = {missing_pma:,}; every blank occurs outside China). Node values are totals within each stage.", ha="left", va="bottom", fontsize=8.5, color="#5D6970")
fig.subplots_adjust(left=0.045, right=0.987, top=0.79, bottom=0.055, wspace=0.10)
output = base / "case29_sankey_python.png"
fig.savefig(output, dpi=360, facecolor=fig.get_facecolor(), bbox_inches="tight", pad_inches=0.12)
plt.close(fig)
print(output)
