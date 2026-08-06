from pathlib import Path
from collections import Counter
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
mpl.rcParams.update({"font.family": "DejaVu Sans", "font.size": 8.5, "axes.linewidth": 0.7})
INK = "#202528"
TEAL = "#287D78"
UP = "#C84A5B"
DOWN = "#4C78A8"
MIX = "#C6923B"

def read(path):
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x) if x is not None else "" for x in rows[0]]
    return h, rows[1:]

def clinical():
    h, rows = read(ROOT / "data1.xlsx")
    sets = h[1:]
    patterns = []
    for row in rows:
        p = tuple(int(float(row[i] or 0) > 0) for i in range(1, len(h)))
        if any(p):
            patterns.append(p)
    return sets, patterns

def timecourse():
    h, rows = read(ROOT / "data2.xlsx")
    idx = {k: i for i, k in enumerate(h)}
    days = ["D1", "D3", "D5", "D7", "D9"]
    sets = [f"{d} Sym" for d in days[::-1]] + [f"{d} Apo" for d in days[::-1]]
    patterns = []
    for row in rows:
        bits_down, bits_up = [], []
        for d in days[::-1]:
            try:
                fc = float(row[idx[f"{d}.log2fc"]])
                q = float(row[idx[f"{d}.padj"]])
            except (TypeError, ValueError):
                fc, q = np.nan, np.nan
            bits_down.append(int(np.isfinite(fc) and np.isfinite(q) and fc < -0.5 and q < 0.05))
            bits_up.append(int(np.isfinite(fc) and np.isfinite(q) and fc > 0.5 and q < 0.05))
        p = tuple(bits_down + bits_up)
        if any(p):
            patterns.append(p)
    return sets, patterns

def panel(fig, slot, sets, patterns, title, letter, time=False):
    inner = slot.subgridspec(2, 1, height_ratios=[1.05, 0.8], hspace=0.04)
    axb = fig.add_subplot(inner[0])
    axm = fig.add_subplot(inner[1], sharex=axb)
    counts = Counter(patterns)
    top = sorted(counts.items(), key=lambda z: (-z[1], z[0]))[:16]
    pats = [x[0] for x in top]
    vals = [x[1] for x in top]
    x = np.arange(len(top))
    colors = []
    for p in pats:
        if not time:
            colors.append(TEAL)
        else:
            lo, hi = any(p[:5]), any(p[5:])
            colors.append(MIX if lo and hi else (UP if hi else DOWN))
    axb.bar(x, vals, color=colors, width=0.72)
    for xi, value in zip(x, vals):
        axb.text(xi, value, str(value), ha="center", va="bottom", fontsize=6.5)
    axb.set_ylabel("Exact intersection size")
    axb.set_title(f"{letter}  {title}", loc="left", fontsize=12, fontweight="bold", color=INK)
    axb.text(0.99, 0.98, f"{len(patterns):,} records in ≥1 set\nTop {len(top)} exact patterns", transform=axb.transAxes, ha="right", va="top", fontsize=7, color="#656B6E")
    axb.spines[["top", "right", "bottom"]].set_visible(False)
    axb.tick_params(axis="x", bottom=False, labelbottom=False)
    axb.grid(axis="y", color="#E7E3DD", lw=0.5)
    for yi in range(len(sets)):
        if yi % 2 == 0:
            axm.axhspan(yi - 0.5, yi + 0.5, color="#F0EEE9", zorder=0)
    for xi, p in enumerate(pats):
        active = np.where(np.array(p) == 1)[0]
        axm.scatter(np.full(len(sets), xi), np.arange(len(sets)), s=12, color="#D6D6D2", zorder=1)
        if len(active):
            axm.plot([xi, xi], [active.min(), active.max()], color=INK, lw=1.2, zorder=2)
            dotcols = []
            for a in active:
                dotcols.append((DOWN if a < 5 else UP) if time else TEAL)
            axm.scatter(np.full(len(active), xi), active, s=26, color=dotcols, zorder=3)
    set_sizes = np.sum(np.array(patterns), axis=0)
    labels = [f"{s}   {n:,}" for s, n in zip(sets, set_sizes)]
    axm.set_yticks(np.arange(len(sets)), labels)
    axm.set_ylim(len(sets) - 0.5, -0.5)
    axm.set_xticks([])
    axm.set_xlabel("Set name  ·  total members")
    axm.spines[["top", "right", "bottom", "left"]].set_visible(False)
    axm.tick_params(axis="y", length=0)

sets1, pat1 = clinical()
sets2, pat2 = timecourse()
fig = plt.figure(figsize=(13.4, 7.8), facecolor="#FBFAF7")
outer = fig.add_gridspec(1, 2, wspace=0.2, left=0.07, right=0.98, bottom=0.1, top=0.87)
panel(fig, outer[0], sets1, pat1, "Clinical comparison intersections", "A", False)
panel(fig, outer[1], sets2, pat2, "Time-course differential intersections", "B", True)
fig.suptitle("Intersection architecture across two studies", x=0.07, y=0.97, ha="left", fontsize=19, fontweight="bold", color=INK)
fig.text(0.07, 0.925, "Exact membership patterns, ranked without collapsing combinations", fontsize=9, color="#656B6E")
fig.text(0.07, 0.025, "Time-course sets require adjusted P < 0.05 and |log₂FC| > 0.5. Sym = negative direction; Apo = positive direction. Clinical sets use supplied binary memberships.", fontsize=7.3, color="#656B6E")
fig.savefig(ROOT / "plot_python.png", dpi=360, facecolor=fig.get_facecolor())
plt.close(fig)
