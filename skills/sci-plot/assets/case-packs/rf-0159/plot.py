from pathlib import Path
from collections import defaultdict, Counter
import json
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon
from matplotlib.collections import PatchCollection
from matplotlib.colors import Normalize, LinearSegmentedColormap
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
mpl.rcParams.update({"font.family": ["Arial Unicode MS", "DejaVu Sans"], "font.size": 8.5, "axes.linewidth": 0.7})
with open(ROOT / "china_city.geojson", encoding="utf-8") as f:
    city_geo = json.load(f)
with open(ROOT / "china_province.geojson", encoding="utf-8") as f:
    province_geo = json.load(f)

def read_xlsx(path):
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x) for x in rows[0]]
    return [dict(zip(h, r)) for r in rows[1:] if any(v is not None for v in r)]

def rings(geom):
    t, c = geom["type"], geom["coordinates"]
    if t == "Polygon":
        return [c[0]]
    if t == "MultiPolygon":
        return [p[0] for p in c]
    return []

def lines(geom):
    t, c = geom["type"], geom["coordinates"]
    if t == "LineString":
        return [c]
    if t == "MultiLineString":
        return c
    return []

def centroid(feature):
    rr = rings(feature["geometry"])
    if not rr:
        ll = lines(feature["geometry"])
        arr = np.array([p for line in ll for p in line], float)
        return tuple(arr.mean(axis=0)) if len(arr) else (np.nan, np.nan)
    ring = max(rr, key=len)
    arr = np.asarray(ring, float)
    x, y = arr[:, 0], arr[:, 1]
    cross = x * np.roll(y, -1) - np.roll(x, -1) * y
    area = cross.sum() / 2
    if abs(area) < 1e-12:
        return float(x.mean()), float(y.mean())
    cx = ((x + np.roll(x, -1)) * cross).sum() / (6 * area)
    cy = ((y + np.roll(y, -1)) * cross).sum() / (6 * area)
    return float(cx), float(cy)

def base(ax, city_fill=None, norm=None, cmap=None):
    patches, values = [], []
    for f in city_geo["features"]:
        name = str(f["properties"].get("name"))
        for ring in rings(f["geometry"]):
            patches.append(Polygon(np.asarray(ring), closed=True))
            values.append(np.nan if city_fill is None else city_fill.get(name, np.nan))
    if city_fill is None:
        coll = PatchCollection(patches, facecolor="#F0EEE7", edgecolor="#FFFFFF", linewidth=0.12, zorder=1)
    else:
        colors = [("#EFEDE6" if not np.isfinite(v) else cmap(norm(v))) for v in values]
        coll = PatchCollection(patches, facecolor=colors, edgecolor="#FFFFFF", linewidth=0.10, zorder=1)
    ax.add_collection(coll)
    for f in province_geo["features"]:
        for ring in rings(f["geometry"]):
            a = np.asarray(ring)
            ax.plot(a[:, 0], a[:, 1], color="#4D5354", lw=0.36, zorder=2)
        for line in lines(f["geometry"]):
            a = np.asarray(line)
            ax.plot(a[:, 0], a[:, 1], color="#4D5354", lw=0.36, zorder=2)
    ax.set_xlim(73, 136)
    ax.set_ylim(17, 54.5)
    ax.set_aspect("equal", adjustable="box")
    ax.set_xticks([])
    ax.set_yticks([])
    ax.spines[:].set_visible(False)

centroids = {str(f["properties"].get("name")): centroid(f) for f in city_geo["features"]}
d1 = read_xlsx(ROOT / "data1.xlsx")
d2 = read_xlsx(ROOT / "data2.xlsx")
by_city = defaultdict(lambda: {"total": 0.0, "genus": Counter()})
for r in d1:
    city = str(r["City"])
    n = float(r["Number of tick samples"] or 0)
    genus = str(r["Tick genus"])
    by_city[city]["total"] += n
    by_city[city]["genus"][genus] += n
genera = sorted({g for z in by_city.values() for g in z["genus"]})
palette = ["#B84C4C", "#D58B3E", "#708C48", "#318884", "#4D75A1", "#7B6599", "#AD617E", "#8B7355"]
gcol = {g: palette[i % len(palette)] for i, g in enumerate(genera)}
value_map = {str(r["City"]): float(r["Value"] or 0) for r in d2}
vals = np.array(list(value_map.values()), float)
norm = Normalize(vmin=0, vmax=np.percentile(vals, 98))
cmap = LinearSegmentedColormap.from_list("nf", ["#F4EADB", "#E9B66C", "#D8674A", "#7B2D45"])
fig, axes = plt.subplots(1, 2, figsize=(14.6, 7.3), facecolor="#FBFAF7")
base(axes[0])
available = [(city, z) for city, z in by_city.items() if city in centroids]
sizes = np.array([z["total"] for _, z in available])
for city, z in available:
    x, y = centroids[city]
    dominant = z["genus"].most_common(1)[0][0]
    s = 10 + 76 * np.sqrt(z["total"] / max(sizes.max(), 1))
    axes[0].scatter(x, y, s=s, c=gcol[dominant], edgecolor="white", linewidth=0.45, alpha=0.86, zorder=4)
axes[0].set_title("A  Tick sampling landscape", loc="left", fontsize=12, fontweight="bold")
axes[0].text(0.01, 0.01, "Bubble area: total tick samples\nColor: dominant tick genus", transform=axes[0].transAxes, fontsize=7, color="#606668", va="bottom")
handles = [plt.Line2D([0], [0], marker="o", linestyle="", markerfacecolor=gcol[g], markeredgecolor="white", label=g, markersize=5.5) for g in genera]
axes[0].legend(handles=handles, title="Dominant genus", loc="lower left", bbox_to_anchor=(0.0, 0.08), ncol=2, frameon=False, fontsize=6.3, title_fontsize=7)
base(axes[1], value_map, norm, cmap)
axes[1].set_title("B  City-level value distribution", loc="left", fontsize=12, fontweight="bold")
sm = mpl.cm.ScalarMappable(norm=norm, cmap=cmap)
cb = fig.colorbar(sm, ax=axes[1], fraction=0.028, pad=0.01, shrink=0.58)
cb.set_label("Value", fontsize=8)
cb.ax.tick_params(labelsize=7)
axes[1].text(0.01, 0.01, "Color scale capped at the 98th percentile\nfor legibility; source values unchanged", transform=axes[1].transAxes, fontsize=7, color="#606668", va="bottom")
fig.suptitle("China spatial evidence atlas", x=0.055, y=0.97, ha="left", fontsize=19, fontweight="bold", color="#202528")
fig.text(0.055, 0.925, "City geometries and province outlines are read directly from the supplied GeoJSON files", fontsize=9, color="#656B6E")
fig.text(0.055, 0.02, "Tick data: 196 of 198 named cities matched exactly; Wanzhou and Fuling districts were not positioned. Administrative boundaries are shown as supplied and imply no endorsement.", fontsize=7.2, color="#656B6E")
fig.tight_layout(rect=(0.025, 0.05, 0.99, 0.9), w_pad=1.5)
fig.savefig(ROOT / "plot_python.png", dpi=360, facecolor=fig.get_facecolor())
plt.close(fig)
